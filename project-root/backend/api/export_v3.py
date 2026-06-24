"""Excel export — 仿宋16号 完全匹配汇总模板格式 + 全部Sheet统一样式"""
import re
from collections import defaultdict
from io import BytesIO
from urllib.parse import quote
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db, get_method_full_name

router = APIRouter(prefix="/export", tags=["export"])

# 模板统一字体：仿宋16号，表头加粗，数据正常
FONT_H = Font(name="仿宋", size=16, bold=True)
FONT_D = Font(name="仿宋", size=16)
ALIGN_C = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin")
BORDER_T = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 严格匹配模板列宽（F修正为19.66，K/M/N对应模板末尾汇总列）
WIDTHS = {'A': 8.89, 'B': 24.89, 'C': 18.0, 'D': 17.44, 'E': 43.66,
          'F': 19.66, 'K': 18.66, 'M': 22.11, 'N': 16.78}

def _cl(n): return get_column_letter(n)

def _auto_fit(ws, min_col=1, max_col=None, data_start=1, tight_cols=None):
    """自动列宽：中文=2宽度，ASCII=1宽度；tight_cols用压缩系数1.2"""
    if max_col is None:
        max_col = ws.max_column
    if tight_cols is None:
        tight_cols = set()
    for col in range(min_col, max_col + 1):
        best = 0
        for row in ws.iter_rows(min_row=data_start, min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    text = str(cell.value)
                    w = sum(2 if '\u4e00' <= c <= '\u9fff' or '\u3000' <= c <= '\u303f'
                            or '\uff00' <= c <= '\uffef' else 1 for c in text)
                    if w > best: best = w
        if col in tight_cols:
            ws.column_dimensions[_cl(col)].width = max(int(best * 1.2) + 2, 5) if best else 6
        else:
            ws.column_dimensions[_cl(col)].width = max(int(best * 1.8) + 6, 10) if best else 10

def _parse_instrument(project_name: str) -> tuple[str, str, str]:
    m = re.match(r'^(.+?)-((?:LC|GC)-\d+)(.*)$', project_name)
    if m:
        prefix, code, suffix = m.group(1), m.group(2), m.group(3)
        itype = "液相" if code.upper().startswith("LC") else "气相"
        return (f"{prefix}-{suffix}" if suffix else prefix, code, itype)
    return (project_name, "", "其他")

def _extract_project_code(n): return n.split("-", 1)[0] if "-" in n else n

def _week_ranges(start: datetime, end: datetime) -> list[tuple[str, datetime, datetime]]:
    weeks = []
    cur = start - timedelta(days=start.weekday())
    while cur <= end:
        wk_end = cur + timedelta(days=6)
        label = f"{cur.strftime('%m.%d')}-{wk_end.strftime('%m.%d')}"
        weeks.append((label, cur, wk_end))
        cur = wk_end + timedelta(days=1)
    return weeks

def _month_bounds(ref_date):
    d_start = ref_date.replace(day=1)
    if d_start.month == 12:
        d_end = d_start.replace(year=d_start.year+1, month=1, day=1) - timedelta(days=1)
    else:
        d_end = d_start.replace(month=d_start.month+1, day=1) - timedelta(days=1)
    return d_start, d_end

def _set_header(ws, row, col, label, font=None):
    """写表头单元格：指定font或默认FONT_H"""
    c = ws.cell(row=row, column=col, value=label)
    c.font = font or FONT_H
    c.alignment = ALIGN_C
    c.border = BORDER_T
    return c

def _set_data(ws, row, col, value):
    """写数据单元格：FONT_D + 居中 + 细线边框"""
    c = ws.cell(row=row, column=col, value=value if value else None)
    c.font = FONT_D
    c.alignment = ALIGN_C
    c.border = BORDER_T
    return c

# ═══════════════════  Sheet 1: 月-汇总 — 严格匹配模板 ═══════════════════
def _build_monthly_summary(ws, db, start_s, end_s, group_id):
    if start_s:
        ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else:
        ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    month_label = f"{d_start.year}年{d_start.month}月"

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None:
        pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT p.id AS project_id, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY p.id", rp).fetchall()
    proj_monthly = {r["project_id"]: (r["qty"] or 0) for r in records}

    lab_order, lab_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        qty = proj_monthly.get(proj["id"], 0)
        if gn not in lab_data:
            lab_data[gn] = {}; lab_order.append(gn)
        if pc_code not in lab_data[gn]:
            lab_data[gn][pc_code] = {'lc': [], 'gc': []}
        if it == "气相":
            lab_data[gn][pc_code]['gc'].append((ic, fn, pn, qty))
        else:
            lab_data[gn][pc_code]['lc'].append((ic, fn, pn, qty))

    # 模板列: B=使用实验室 C=项目代号 D=液相仪器 E=检测方法
    #         F=月检测数量 G=液相检测量 H=气相检测量 I=项目检测总量
    CB, CC, CD, CE = 2, 3, 4, 5
    CF, CG, CH, CI = 6, 7, 8, 9
    HR = 2

    # 表头 — 全部 FONT_H bold
    for col, label in [(CB, "使用实验室"), (CC, "项目代号"), (CD, "液相仪器"), (CE, "检测方法"),
                        (CF, "月检测数量"), (CG, "液相检测量"), (CH, "气相检测量"), (CI, "项目检测总量")]:
        _set_header(ws, HR, col, label)

    # 数据行 — 扁平结构，无小计行，仅合并单元格
    r = HR + 1
    for lab_name in lab_order:
        pd = lab_data[lab_name]
        lab_start = r
        for pc_code, inst in pd.items():
            psr = r
            for ic, ml, pn, qty in inst['lc'] + inst['gc']:
                _set_data(ws, r, CB, lab_name)
                _set_data(ws, r, CC, pc_code)
                _set_data(ws, r, CD, ic)
                _set_data(ws, r, CE, ml)
                _set_data(ws, r, CF, qty if qty else None)
                r += 1
            per = r - 1
            if psr > per: continue
            # 合并项目代号列
            if per > psr:
                ws.merge_cells(start_row=psr, start_column=CC, end_row=per, end_column=CC)
            # 合并 G/H/I 列（项目级别汇总列）
            for col in [CG, CH, CI]:
                if per > psr:
                    ws.merge_cells(start_row=psr, start_column=col, end_row=per, end_column=col)

            # 液相/气相检测量 = SUM(各自在F列的范围)
            lc_start = psr
            lc_end = psr + len(inst['lc']) - 1 if inst['lc'] else psr - 1
            gc_start = psr + len(inst['lc'])
            gc_end = per if inst['gc'] else gc_start - 1

            if inst['lc'] and inst['gc']:
                ws.cell(row=psr, column=CG).value = f"=SUM({_cl(CF)}{lc_start}:{_cl(CF)}{lc_end})"
                ws.cell(row=psr, column=CH).value = f"=SUM({_cl(CF)}{gc_start}:{_cl(CF)}{gc_end})"
            elif inst['lc']:
                ws.cell(row=psr, column=CG).value = f"=SUM({_cl(CF)}{lc_start}:{_cl(CF)}{lc_end})"
            elif inst['gc']:
                ws.cell(row=psr, column=CH).value = f"=SUM({_cl(CF)}{gc_start}:{_cl(CF)}{gc_end})"

            # 项目检测总量 = 液相 + 气相
            ws.cell(row=psr, column=CI).value = f"=SUM({_cl(CG)}{psr}:{_cl(CH)}{per})"

            # 给 G/H/I 合并区域所有行刷格式
            for col in [CG, CH, CI]:
                for br in range(psr, per + 1):
                    ws.cell(row=br, column=col).font = FONT_D
                    ws.cell(row=br, column=col).alignment = ALIGN_C
                    ws.cell(row=br, column=col).border = BORDER_T

        # 合并实验室列
        ler = r - 1
        if ler > lab_start:
            ws.merge_cells(start_row=lab_start, start_column=CB, end_row=ler, end_column=CB)

    # 总计行 — 模板为 normal 字体（非bold）
    total_row = r
    _set_data(ws, total_row, CB, "总计")
    _set_data(ws, total_row, CC, "")
    _set_data(ws, total_row, CD, "")
    _set_data(ws, total_row, CE, "")
    ws.cell(row=total_row, column=CF, value=f"=SUM({_cl(CF)}{HR + 1}:{_cl(CF)}{total_row - 1})")
    ws.cell(row=total_row, column=CG, value=f"=SUM({_cl(CG)}{HR + 1}:{_cl(CG)}{total_row - 1})")
    ws.cell(row=total_row, column=CH, value=f"=SUM({_cl(CH)}{HR + 1}:{_cl(CH)}{total_row - 1})")
    ws.cell(row=total_row, column=CI, value=f"=SUM({_cl(CI)}{HR + 1}:{_cl(CI)}{total_row - 1})")
    for col in [CF, CG, CH, CI]:
        ws.cell(row=total_row, column=col).font = FONT_D
        ws.cell(row=total_row, column=col).alignment = ALIGN_C
        ws.cell(row=total_row, column=col).border = BORDER_T

    # 模板列宽（WIDTHS覆盖auto_fit的通用计算）
    _auto_fit(ws, min_col=CB, max_col=CI, data_start=HR)
    for col_letter, w in WIDTHS.items():
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = f"{_cl(CE)}{HR + 1}"
    return month_label


# ═══════════════════  Sheet 2: 每日工作量 — 模板统一样式 ═══════════════════
def _build_daily_work(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "1976D2"
    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT date(wr.recorded_at) AS work_day, pg.name AS group_name, p.name AS project_name, SUM(wr.quantity) AS qty, GROUP_CONCAT(DISTINCT wr.user_name) AS users FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY date(wr.recorded_at), pg.id, p.id ORDER BY work_day, pg.sort_order, p.sort_order", rp).fetchall()

    hd = ["日期", "实验室", "项目代码", "仪器", "检测方法", "类型", "数量", "录入人"]
    for ci, lb in enumerate(hd, 1):
        _set_header(ws, 1, ci, lb)

    for ri, row in enumerate(rows, 2):
        gn, pn = row["group_name"], row["project_name"]
        pc = _extract_project_code(pn); mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        wd = row["work_day"]; wdd = wd if isinstance(wd, str) else str(wd)
        vals = [wdd, gn, pc, ic, fn, it, row["qty"], row["users"] or ""]
        for ci, val in enumerate(vals, 1):
            _set_data(ws, ri, ci, val)

    _auto_fit(ws)
    ws.freeze_panes = "A2"


# ═══════════════════  Sheet 3: 每周工作量 — 严格匹配模板格式 ═══════════════════
def _build_weekly_work(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "43A047"
    if start_s:
        ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else:
        ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    weeks = _week_ranges(d_start, d_end)
    num_weeks = len(weeks)

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None:
        pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT p.id AS project_id, date(wr.recorded_at) AS work_day, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY p.id, date(wr.recorded_at)", rp).fetchall()

    proj_weekly = defaultdict(lambda: defaultdict(int))
    for row in records:
        rd = datetime.strptime(row["work_day"], "%Y-%m-%d")
        for wi, (_, ws_d, we_d) in enumerate(weeks):
            if ws_d <= rd <= we_d:
                proj_weekly[row["project_id"]][wi] += (row["qty"] or 0)
                break

    lab_order, lab_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        dm = proj_weekly.get(proj["id"], {})
        if gn not in lab_data: lab_data[gn] = {}; lab_order.append(gn)
        if pc_code not in lab_data[gn]: lab_data[gn][pc_code] = {'lc': [], 'gc': []}
        if it == "气相":
            lab_data[gn][pc_code]['gc'].append((ic, fn, dict(dm)))
        else:
            lab_data[gn][pc_code]['lc'].append((ic, fn, dict(dm)))

    # 模板列: B-E = 使用实验室/项目代号/液相仪器/检测方法
    #         F-J = 合并表头 "周汇总（X月）"（F起点）
    #         K=液相检测量, L=气相检测量, M=项目检测总量
    CB, CC, CD, CE = 2, 3, 4, 5
    CW = 6                       # 周数据起始列 F
    CW_END = CW + num_weeks - 1  # 周数据结束列
    CK, CL, CM = CW + num_weeks, CW + num_weeks + 1, CW + num_weeks + 2
    LAST = CM
    HR = 2

    # 表头行 — 完全匹配模板
    for col, label in [(CB, "使用实验室"), (CC, "项目代号"), (CD, "液相仪器"), (CE, "检测方法")]:
        _set_header(ws, HR, col, label)
    for col, label in [(CK, "液相检测量"), (CL, "气相检测量"), (CM, "项目检测总量")]:
        _set_header(ws, HR, col, label)

    # F-J 合并为 "周汇总（X月）"
    if num_weeks > 0:
        ws.merge_cells(start_row=HR, start_column=CW, end_row=HR, end_column=CW_END)
        c = ws.cell(row=HR, column=CW, value=f"周汇总（{d_start.month}月）")
        c.font = FONT_H; c.alignment = ALIGN_C
        for dc in range(CW, CW_END + 1):
            ws.cell(row=HR, column=dc).border = BORDER_T

    # 数据行 — 与月-汇总相同结构和合并规则
    # 区别：F-J填每周数值，K=SUM(LC周的F), L=SUM(GC周的F), M=SUM(K:L)
    r = HR + 1
    for lab_name in lab_order:
        pd = lab_data[lab_name]
        lab_start = r
        for pc_code, inst in pd.items():
            psr = r
            for ic, ml, dm in inst['lc'] + inst['gc']:
                _set_data(ws, r, CB, lab_name)
                _set_data(ws, r, CC, pc_code)
                _set_data(ws, r, CD, ic)
                _set_data(ws, r, CE, ml)
                for wi in range(num_weeks):
                    qty = dm.get(wi, 0)
                    _set_data(ws, r, CW + wi, qty)
                r += 1
            per = r - 1
            if psr > per: continue
            if per > psr:
                ws.merge_cells(start_row=psr, start_column=CC, end_row=per, end_column=CC)
            for col in [CK, CL, CM]:
                if per > psr:
                    ws.merge_cells(start_row=psr, start_column=col, end_row=per, end_column=col)

            # 液相/气相检测量 = SUM(各自每周行范围)
            lc_end_row = psr + len(inst['lc']) - 1 if inst['lc'] else psr - 1
            gc_start_row = psr + len(inst['lc'])
            gc_end_row = per if inst['gc'] else gc_start_row - 1

            if inst['lc'] and inst['gc']:
                lc_parts = ",".join(f"{_cl(CW + wi)}{psr}:{_cl(CW + wi)}{lc_end_row}" for wi in range(num_weeks))
                gc_parts = ",".join(f"{_cl(CW + wi)}{gc_start_row}:{_cl(CW + wi)}{gc_end_row}" for wi in range(num_weeks))
                ws.cell(row=psr, column=CK).value = f"=SUM({lc_parts})" if lc_end_row >= psr else 0
                ws.cell(row=psr, column=CL).value = f"=SUM({gc_parts})" if gc_end_row >= gc_start_row else 0
            elif inst['lc']:
                lc_parts = ",".join(f"{_cl(CW + wi)}{psr}:{_cl(CW + wi)}{lc_end_row}" for wi in range(num_weeks))
                ws.cell(row=psr, column=CK).value = f"=SUM({lc_parts})"
            elif inst['gc']:
                gc_parts = ",".join(f"{_cl(CW + wi)}{gc_start_row}:{_cl(CW + wi)}{gc_end_row}" for wi in range(num_weeks))
                ws.cell(row=psr, column=CL).value = f"=SUM({gc_parts})"

            # 项目检测总量
            ws.cell(row=psr, column=CM).value = f"=SUM({_cl(CK)}{psr}:{_cl(CL)}{per})"

            # 刷汇总列格式
            for col in [CK, CL, CM]:
                for br in range(psr, per + 1):
                    ws.cell(row=br, column=col).font = FONT_D
                    ws.cell(row=br, column=col).alignment = ALIGN_C
                    ws.cell(row=br, column=col).border = BORDER_T

        ler = r - 1
        if ler > lab_start:
            ws.merge_cells(start_row=lab_start, start_column=CB, end_row=ler, end_column=CB)

    # 总计行 — 模板 normal 字体
    total_row = r
    _set_data(ws, total_row, CB, "总计")
    _set_data(ws, total_row, CC, "")
    _set_data(ws, total_row, CD, "")
    _set_data(ws, total_row, CE, "")
    for wi in range(num_weeks):
        cl = _cl(CW + wi)
        ws.cell(row=total_row, column=CW + wi, value=f"=SUM({cl}{HR + 1}:{cl}{total_row - 1})")
        ws.cell(row=total_row, column=CW + wi).font = FONT_D
        ws.cell(row=total_row, column=CW + wi).alignment = ALIGN_C
        ws.cell(row=total_row, column=CW + wi).border = BORDER_T
    for col, label in [(CK, f"=SUM({_cl(CK)}{HR + 1}:{_cl(CK)}{total_row - 1})"),
                        (CL, f"=SUM({_cl(CL)}{HR + 1}:{_cl(CL)}{total_row - 1})"),
                        (CM, f"=SUM({_cl(CM)}{HR + 1}:{_cl(CM)}{total_row - 1})")]:
        ws.cell(row=total_row, column=col, value=label)
        ws.cell(row=total_row, column=col).font = FONT_D
        ws.cell(row=total_row, column=col).alignment = ALIGN_C
        ws.cell(row=total_row, column=col).border = BORDER_T

    # 列宽：先自动计算（不包括模板固定列），再覆盖模板精确值
    for wi in range(num_weeks):
        ws.column_dimensions[_cl(CW + wi)].width = 8
    _auto_fit(ws, min_col=CB, max_col=LAST, data_start=HR,
              tight_cols=set(range(CW, CW + num_weeks)))
    for col_letter, w in WIDTHS.items():
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = f"{_cl(CE)}{HR + 1}"


# ═══════════════════  Sheet 4: 原始记录 — 模板统一样式 ═══════════════════
def _build_raw_records(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "FF9800"
    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT pg.name AS group_name, p.name AS project_name, wr.user_name, wr.quantity, wr.recorded_at FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} ORDER BY wr.recorded_at, pg.sort_order, p.sort_order", rp).fetchall()

    hd = ["序号", "日期", "实验室", "项目名称", "仪器", "检测方法", "仪器类型", "数量", "录入人"]
    for ci, lb in enumerate(hd, 1):
        _set_header(ws, 1, ci, lb)

    for ri, row in enumerate(rows, 2):
        gn, pn = row["group_name"], row["project_name"]
        mb, ic, it = _parse_instrument(pn)
        rd = row["recorded_at"]; rdd = rd[:10] if isinstance(rd, str) else str(rd)[:10]
        vals = [ri - 1, rdd, gn, pn, ic, mb, it, row["quantity"], row["user_name"]]
        for ci, val in enumerate(vals, 1):
            _set_data(ws, ri, ci, val)

    _auto_fit(ws)
    ws.freeze_panes = "A2"


# ═══════════════════  Sheet 5: 用户统计 — 模板统一样式（按用户分组+合并） ═══════════════════
def _build_user_stats(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "9C27B0"
    if start_s: ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else: ref = datetime.now()
    d_start, d_end = _month_bounds(ref)

    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT wr.user_name, pg.name AS group_name, p.name AS project_name, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY wr.user_name, p.id ORDER BY wr.user_name, pg.sort_order, p.sort_order", rp).fetchall()

    # 模板列: B=用户名 C=使用实验室 D=项目代号 E=液相仪器 F=检测方法 G=数量
    CB, CC, CD, CE, CF, CG = 2, 3, 4, 5, 6, 7
    HR = 2

    for col, label in [(CB, "用户名"), (CC, "使用实验室"), (CD, "项目代号"), (CE, "液相仪器"), (CF, "检测方法"), (CG, "数量")]:
        _set_header(ws, HR, col, label)

    # 按用户分组，用户名列合并
    r = HR + 1
    current_user = None
    user_start = None

    def _finish_user_merge():
        nonlocal current_user, user_start
        if current_user and r - 1 > user_start:
            ws.merge_cells(start_row=user_start, start_column=CB, end_row=r - 1, end_column=CB)

    for row in rows:
        un, gn, pn = row["user_name"], row["group_name"], row["project_name"]
        pc = _extract_project_code(pn); mb, ic, _ = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb

        if un != current_user:
            _finish_user_merge()
            current_user = un
            user_start = r

        _set_data(ws, r, CB, un)
        _set_data(ws, r, CC, gn)
        _set_data(ws, r, CD, pc)
        _set_data(ws, r, CE, ic)
        _set_data(ws, r, CF, fn)
        _set_data(ws, r, CG, row["qty"] or None)
        r += 1

    _finish_user_merge()

    ws.column_dimensions['A'].width = 3
    _auto_fit(ws, min_col=CB, max_col=CG, data_start=HR)
    ws.freeze_panes = f"{_cl(CF)}{HR + 1}"


# ═══════════════════  导出入口 ═══════════════════
@router.get("/excel")
def export_excel(start: str = Query(None), end: str = Query(None), group_id: int = Query(None), db=Depends(get_db)):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "月-汇总"; ws1.sheet_properties.tabColor = "1976D2"
    month_label = _build_monthly_summary(ws1, db, start, end, group_id)
    ws2 = wb.create_sheet(title="每日工作量"); _build_daily_work(ws2, db, start, end, group_id)
    ws3 = wb.create_sheet(title="每周工作量"); _build_weekly_work(ws3, db, start, end, group_id)
    ws4 = wb.create_sheet(title="原始记录"); _build_raw_records(ws4, db, start, end, group_id)
    ws5 = wb.create_sheet(title="用户统计"); _build_user_stats(ws5, db, start, end, group_id)
    output = BytesIO(); wb.save(output); output.seek(0)
    filename = quote(f"工作量统计_{month_label}.xlsx")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
