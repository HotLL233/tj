"""Excel export — 全部Sheet统一月-汇总模板格式（仿宋16号）"""
import re
from collections import defaultdict
from io import BytesIO
from urllib.parse import quote
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db, get_method_full_name

router = APIRouter(prefix="/export", tags=["export"])

FONT_H = Font(name="仿宋", size=16, bold=True)
FONT_D = Font(name="仿宋", size=16)
ALIGN_C = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin")
BORDER_T = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 模板精确列宽
WIDTHS = {'A': 8.89, 'B': 24.89, 'C': 18.0, 'D': 17.44, 'E': 43.66,
          'F': 19.66, 'K': 18.66, 'M': 22.11, 'N': 16.78}

def _cl(n): return get_column_letter(n)

def _auto_fit(ws, min_col=1, max_col=None, data_start=1, tight_cols=None):
    if max_col is None: max_col = ws.max_column
    if tight_cols is None: tight_cols = set()
    for col in range(min_col, max_col + 1):
        best = 0
        for row in ws.iter_rows(min_row=data_start, min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    text = str(cell.value)
                    w = sum(2 if '\u4e00' <= c <= '\u9fff' or '\u3000' <= c <= '\u303f'
                            or '\uff00' <= c <= '\uffef' else 1 for c in text)
                    if w > best: best = w
        if col in tight_cols:
            ws.column_dimensions[_cl(col)].width = max(int(best * 1.2) + 2, 5) if best else 6
        else:
            ws.column_dimensions[_cl(col)].width = max(int(best * 1.8) + 6, 10) if best else 10

def _parse_instrument(project_name: str) -> tuple[str, str, str]:
    m = re.match(r'^(.+?)-((?:LC|GC)-\d+)(.*)$', project_name)
    if m:
        prefix, code, suffix = m.group(1), m.group(2), m.group(3)
        itype = "液相" if code.upper().startswith("LC") else "气相"
        return (f"{prefix}-{suffix}" if suffix else prefix, code, itype)
    return (project_name, "", "其他")

def _extract_project_code(n): return n.split("-", 1)[0] if "-" in n else n

def _week_ranges(start: datetime, end: datetime) -> list[tuple[str, datetime, datetime]]:
    weeks = []
    cur = start - timedelta(days=start.weekday())
    while cur <= end:
        wk_end = cur + timedelta(days=6)
        label = f"{cur.strftime('%m.%d')}-{wk_end.strftime('%m.%d')}"
        weeks.append((label, cur, wk_end))
        cur = wk_end + timedelta(days=1)
    return weeks

def _month_bounds(ref_date):
    d_start = ref_date.replace(day=1)
    if d_start.month == 12:
        d_end = d_start.replace(year=d_start.year+1, month=1, day=1) - timedelta(days=1)
    else:
        d_end = d_start.replace(month=d_start.month+1, day=1) - timedelta(days=1)
    return d_start, d_end

def _set_header(ws, row, col, label):
    c = ws.cell(row=row, column=col, value=label)
    c.font = FONT_H; c.alignment = ALIGN_C; c.border = BORDER_T

def _set_data(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value if value else None)
    c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T


# ═══════════════════  公共：按模板格式写出三级树形数据 + 合并 + 公式 ═══════════════════
# _build_tree_rows 返回 row_data 列表，每个元素:
#   (一级标签, 二级标签, 仪器, 方法, 单行数量, is_gc_bool)
# 写入时：一级合并B列、二级合并C列、G/H/I合并并填SUM公式、总计行
_B = 2; _C = 3; _D = 4; _E = 5; _F = 6; _G = 7; _H = 8; _I = 9
_HR = 2

def _write_template_sheet(ws, headers, row_data, extra_cols=None):
    """row_data: [(level1, level2, instrument, method, qty, is_gc), ...]
       按 level1→level2→instrument 层级写模板格式数据。
       extra_cols: {col: value} dict to write AFTER column I (e.g. weekly breakdowns).
       返回 (总行数所在行号, 数据起始行)。
    """
    # 模板固定表头
    for col, label in zip([_B, _C, _D, _E, _F, _G, _H, _I], headers):
        _set_header(ws, _HR, col, label)

    # 附加列表头（周数据等）
    if extra_cols:
        for col, label in extra_cols.items():
            _set_header(ws, _HR, col, label)

    r = _HR + 1
    level1_start = None
    current_l1 = None

    for i, (l1, l2, instr, method, qty, is_gc) in enumerate(row_data):
        # 一级标签变更 → 合并上一级
        if l1 != current_l1:
            if current_l1 and r - 1 > level1_start:
                ws.merge_cells(start_row=level1_start, start_column=_B, end_row=r - 1, end_column=_B)
            current_l1 = l1
            level1_start = r

        _set_data(ws, r, _B, l1)
        _set_data(ws, r, _C, l2)
        _set_data(ws, r, _D, instr)
        _set_data(ws, r, _E, method)
        _set_data(ws, r, _F, qty if qty else None)
        r += 1

    # 合并最后一级
    if current_l1 and r - 1 > level1_start:
        ws.merge_cells(start_row=level1_start, start_column=_B, end_row=r - 1, end_column=_B)

    # 合并二级标签（C列）：连续相同 l2 的合并
    ds = _HR + 1
    while ds < r:
        l2_val = ws.cell(row=ds, column=_C).value
        de = ds
        while de + 1 < r and ws.cell(row=de + 1, column=_C).value == l2_val:
            de += 1
        if de > ds:
            ws.merge_cells(start_row=ds, start_column=_C, end_row=de, end_column=_C)
        ds = de + 1

    # 二级分组（C列合并区域）写 G/H/I 公式和合并
    c_start = _HR + 1
    while c_start < r:
        # 找这个C合并区域的起止
        c_end = c_start
        for check in range(c_start + 1, r):
            c_val = ws.cell(row=check, column=_C).value
            if c_val is not None and c_val != ws.cell(row=c_start, column=_C).value:
                break
            c_end = check

        if c_end >= c_start:
            # 合并 G/H/I
            for col in [_G, _H, _I]:
                if c_end > c_start:
                    ws.merge_cells(start_row=c_start, start_column=col, end_row=c_end, end_column=col)

            # 统计该二级分组内 LC 和 GC 的行范围
            lc_rows, gc_rows = [], []
            for ri in range(c_start, c_end + 1):
                if row_data[ri - (_HR + 1)][5]:  # is_gc
                    gc_rows.append(ri)
                else:
                    lc_rows.append(ri)

            if lc_rows and gc_rows:
                ws.cell(row=c_start, column=_G).value = f"=SUM({_cl(_F)}{min(lc_rows)}:{_cl(_F)}{max(lc_rows)})"
                ws.cell(row=c_start, column=_H).value = f"=SUM({_cl(_F)}{min(gc_rows)}:{_cl(_F)}{max(gc_rows)})"
            elif lc_rows:
                ws.cell(row=c_start, column=_G).value = f"=SUM({_cl(_F)}{min(lc_rows)}:{_cl(_F)}{max(lc_rows)})"
            elif gc_rows:
                ws.cell(row=c_start, column=_H).value = f"=SUM({_cl(_F)}{min(gc_rows)}:{_cl(_F)}{max(gc_rows)})"

            ws.cell(row=c_start, column=_I).value = f"=SUM({_cl(_G)}{c_start}:{_cl(_H)}{c_end})"

            # 刷 G/H/I 格式
            for col in [_G, _H, _I]:
                for ri in range(c_start, c_end + 1):
                    ws.cell(row=ri, column=col).font = FONT_D
                    ws.cell(row=ri, column=col).alignment = ALIGN_C
                    ws.cell(row=ri, column=col).border = BORDER_T

        c_start = c_end + 1

    # 附加列数据（周数据等，每行独立）
    if extra_cols:
        for col in extra_cols:
            for ri in range(_HR + 1, r):
                _set_data(ws, ri, col, None)  # 预填空+格式（由调用方后续填充）

    # 总计行
    total_row = r
    _set_data(ws, total_row, _B, "总计")
    _set_data(ws, total_row, _C, "")
    _set_data(ws, total_row, _D, "")
    _set_data(ws, total_row, _E, "")
    sum_range = f"{_HR + 1}:{total_row - 1}"
    for col in [_F, _G, _H, _I]:
        ws.cell(row=total_row, column=col, value=f"=SUM({_cl(col)}{sum_range})")
        ws.cell(row=total_row, column=col).font = FONT_D
        ws.cell(row=total_row, column=col).alignment = ALIGN_C
        ws.cell(row=total_row, column=col).border = BORDER_T
    if extra_cols:
        for col in extra_cols:
            ws.cell(row=total_row, column=col, value=f"=SUM({_cl(col)}{sum_range})")
            ws.cell(row=total_row, column=col).font = FONT_D
            ws.cell(row=total_row, column=col).alignment = ALIGN_C
            ws.cell(row=total_row, column=col).border = BORDER_T

    return total_row, _HR + 1


# ═══════════════════  Sheet 1: 月-汇总 — 严格匹配模板 ═══════════════════
def _build_monthly_summary(ws, db, start_s, end_s, group_id):
    if start_s: ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else: ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    month_label = f"{d_start.year}年{d_start.month}月"

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None: pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT p.id AS project_id, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY p.id", rp).fetchall()
    proj_monthly = {r["project_id"]: (r["qty"] or 0) for r in records}

    lab_order, lab_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        qty = proj_monthly.get(proj["id"], 0)
        if gn not in lab_data: lab_data[gn] = {}; lab_order.append(gn)
        if pc_code not in lab_data[gn]: lab_data[gn][pc_code] = {'lc': [], 'gc': []}
        if it == "气相": lab_data[gn][pc_code]['gc'].append((ic, fn, qty))
        else: lab_data[gn][pc_code]['lc'].append((ic, fn, qty))

    row_data = []
    for lab_name in lab_order:
        for pc_code, inst in lab_data[lab_name].items():
            for ic, ml, qty in inst['lc']:
                row_data.append((lab_name, pc_code, ic, ml, qty, False))
            for ic, ml, qty in inst['gc']:
                row_data.append((lab_name, pc_code, ic, ml, qty, True))

    headers = ["使用实验室", "项目代号", "液相仪器", "检测方法", "月检测数量", "液相检测量", "气相检测量", "项目检测总量"]
    _write_template_sheet(ws, headers, row_data)

    _auto_fit(ws, min_col=_B, max_col=_I, data_start=_HR)
    for cl, w in WIDTHS.items(): ws.column_dimensions[cl].width = w
    ws.freeze_panes = f"{_cl(_E)}{_HR + 1}"
    return month_label


# ═══════════════════  Sheet 2: 每日工作量 — 模板格式（4级：日期→实验室→代号→仪器） ═══════════════════
def _build_daily_work(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "1976D2"
    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT date(wr.recorded_at) AS work_day, pg.name AS group_name, p.name AS project_name, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY date(wr.recorded_at), p.id ORDER BY work_day, pg.sort_order, p.sort_order", rp).fetchall()

    # 4级结构: date → lab → code → {lc:[], gc:[]}
    date_order, date_data = [], {}
    for row in rows:
        wd = row["work_day"]; wdd = wd if isinstance(wd, str) else str(wd)
        gn, pn = row["group_name"], row["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        if wdd not in date_data: date_data[wdd] = ([], {}); date_order.append(wdd)
        if gn not in date_data[wdd][1]: date_data[wdd][1][gn] = ([], {}); date_data[wdd][0].append(gn)
        if pc_code not in date_data[wdd][1][gn][1]:
            date_data[wdd][1][gn][1][pc_code] = {'lc': [], 'gc': []}
            date_data[wdd][1][gn][0].append(pc_code)
        if it == "气相": date_data[wdd][1][gn][1][pc_code]['gc'].append((ic, fn, row["qty"] or 0))
        else: date_data[wdd][1][gn][1][pc_code]['lc'].append((ic, fn, row["qty"] or 0))

    # 表头: B=日期 C=使用实验室 D=项目代号 E=检测方法 F=日检测数量 G=液相 H=气相 I=总量
    CB, CC, CD, CE = _B, _C, _D, _E
    CF, CG, CH, CI = _F, _G, _H, _I
    for col, lb in [(CB, "日期"), (CC, "使用实验室"), (CD, "项目代号"), (CE, "检测方法"),
                     (CF, "日检测数量"), (CG, "液相检测量"), (CH, "气相检测量"), (CI, "项目检测总量")]:
        _set_header(ws, _HR, col, lb)

    r = _HR + 1
    for wdd in date_order:
        date_start = r
        for gn in date_data[wdd][0]:
            lab_start = r
            for pc_code in date_data[wdd][1][gn][0]:
                inst = date_data[wdd][1][gn][1][pc_code]
                code_start = r
                for ic, ml, qty in inst['lc'] + inst['gc']:
                    _set_data(ws, r, CB, wdd)
                    _set_data(ws, r, CC, gn)
                    _set_data(ws, r, CD, pc_code)
                    _set_data(ws, r, CE, ml)
                    _set_data(ws, r, CF, qty if qty else None)
                    r += 1
                code_end = r - 1
                if code_end > code_start:
                    ws.merge_cells(start_row=code_start, start_column=CD, end_row=code_end, end_column=CD)
                    for col in [CG, CH, CI]:
                        ws.merge_cells(start_row=code_start, start_column=col, end_row=code_end, end_column=col)
                # G/H/I 公式（同月汇总逻辑）
                lc_end = code_start + len(inst['lc']) - 1 if inst['lc'] else code_start - 1
                gc_start = code_start + len(inst['lc'])
                gc_end = code_end if inst['gc'] else gc_start - 1
                if inst['lc'] and inst['gc']:
                    ws.cell(row=code_start, column=CG).value = f"=SUM({_cl(CF)}{code_start}:{_cl(CF)}{lc_end})"
                    ws.cell(row=code_start, column=CH).value = f"=SUM({_cl(CF)}{gc_start}:{_cl(CF)}{gc_end})"
                elif inst['lc']:
                    ws.cell(row=code_start, column=CG).value = f"=SUM({_cl(CF)}{code_start}:{_cl(CF)}{lc_end})"
                elif inst['gc']:
                    ws.cell(row=code_start, column=CH).value = f"=SUM({_cl(CF)}{gc_start}:{_cl(CF)}{gc_end})"
                ws.cell(row=code_start, column=CI).value = f"=SUM({_cl(CG)}{code_start}:{_cl(CH)}{code_end})"
                for col in [CG, CH, CI]:
                    for br in range(code_start, code_end + 1):
                        ws.cell(row=br, column=col).font = FONT_D
                        ws.cell(row=br, column=col).alignment = ALIGN_C
                        ws.cell(row=br, column=col).border = BORDER_T
            # 合并实验室列
            lab_end = r - 1
            if lab_end > lab_start:
                ws.merge_cells(start_row=lab_start, start_column=CC, end_row=lab_end, end_column=CC)
        # 合并日期列
        date_end = r - 1
        if date_end > date_start:
            ws.merge_cells(start_row=date_start, start_column=CB, end_row=date_end, end_column=CB)

    # 总计行
    total_row = r
    _set_data(ws, total_row, CB, "总计")
    _set_data(ws, total_row, CC, "")
    _set_data(ws, total_row, CD, "")
    _set_data(ws, total_row, CE, "")
    sr = f"{_HR + 1}:{total_row - 1}"
    for col in [CF, CG, CH, CI]:
        ws.cell(row=total_row, column=col, value=f"=SUM({_cl(col)}{sr})")
        ws.cell(row=total_row, column=col).font = FONT_D
        ws.cell(row=total_row, column=col).alignment = ALIGN_C
        ws.cell(row=total_row, column=col).border = BORDER_T

    _auto_fit(ws, min_col=_B, max_col=_I, data_start=_HR)
    for cl, w in WIDTHS.items(): ws.column_dimensions[cl].width = w
    ws.freeze_panes = f"{_cl(CE)}{_HR + 1}"


# ═══════════════════  Sheet 3: 每周工作量 — 和月一样的模板 + 周数据列 ═══════════════════
def _build_weekly_work(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "43A047"
    if start_s: ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else: ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    weeks = _week_ranges(d_start, d_end)
    num_weeks = len(weeks)

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None: pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT p.id AS project_id, date(wr.recorded_at) AS work_day, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY p.id, date(wr.recorded_at)", rp).fetchall()

    proj_weekly = defaultdict(lambda: defaultdict(int))
    for row in records:
        rd = datetime.strptime(row["work_day"], "%Y-%m-%d")
        for wi, (_, ws_d, we_d) in enumerate(weeks):
            if ws_d <= rd <= we_d:
                proj_weekly[row["project_id"]][wi] += (row["qty"] or 0)
                break

    lab_order, lab_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        dm = proj_weekly.get(proj["id"], {})
        if gn not in lab_data: lab_data[gn] = {}; lab_order.append(gn)
        if pc_code not in lab_data[gn]: lab_data[gn][pc_code] = {'lc': [], 'gc': []}
        if it == "气相": lab_data[gn][pc_code]['gc'].append((ic, fn, dict(dm)))
        else: lab_data[gn][pc_code]['lc'].append((ic, fn, dict(dm)))

    # 先按 lab→code→instrument 排好 row_data，qty 填 F列（=所有周之和）
    row_data = []
    for lab_name in lab_order:
        for pc_code, inst in lab_data[lab_name].items():
            for ic, ml, dm in inst['lc']:
                row_data.append((lab_name, pc_code, ic, ml, sum(dm.values()), False))
            for ic, ml, dm in inst['gc']:
                row_data.append((lab_name, pc_code, ic, ml, sum(dm.values()), True))

    # extra_cols: J=第1周, K=第2周, ... 放在 I 后面
    extra = {}
    for wi, (label, _, _) in enumerate(weeks):
        extra[_I + 1 + wi] = label  # J, K, L, ...

    headers = ["使用实验室", "项目代号", "液相仪器", "检测方法", "周检测数量", "液相检测量", "气相检测量", "项目检测总量"]
    total_row, data_start = _write_template_sheet(ws, headers, row_data, extra)

    # 填每周数值（从 J 列开始）
    ri = data_start
    for lab_name in lab_order:
        for pc_code, inst in lab_data[lab_name].items():
            for ic, ml, dm in inst['lc']:
                for wi in range(num_weeks):
                    _set_data(ws, ri, _I + 1 + wi, dm.get(wi, 0))
                ri += 1
            for ic, ml, dm in inst['gc']:
                for wi in range(num_weeks):
                    _set_data(ws, ri, _I + 1 + wi, dm.get(wi, 0))
                ri += 1

    # 列宽：周列固定 8，其余 auto_fit + WIDTHS
    for wi in range(num_weeks):
        ws.column_dimensions[_cl(_I + 1 + wi)].width = 8
    last_col = _I + num_weeks
    _auto_fit(ws, min_col=_B, max_col=last_col, data_start=_HR,
              tight_cols=set(range(_I + 1, _I + 1 + num_weeks)))
    for cl, w in WIDTHS.items(): ws.column_dimensions[cl].width = w
    ws.freeze_panes = f"{_cl(_E)}{_HR + 1}"


# ═══════════════════  Sheet 4: 原始记录 ═══════════════════
def _build_raw_records(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "FF9800"
    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT pg.name AS group_name, p.name AS project_name, wr.user_name, wr.quantity, wr.recorded_at FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} ORDER BY wr.recorded_at, pg.sort_order, p.sort_order", rp).fetchall()

    hd = ["序号", "日期", "实验室", "项目名称", "仪器", "检测方法", "仪器类型", "数量", "录入人"]
    for ci, lb in enumerate(hd, 1): _set_header(ws, 1, ci, lb)
    for ri, row in enumerate(rows, 2):
        gn, pn = row["group_name"], row["project_name"]
        mb, ic, it = _parse_instrument(pn)
        rd = row["recorded_at"]; rdd = rd[:10] if isinstance(rd, str) else str(rd)[:10]
        for ci, val in enumerate([ri - 1, rdd, gn, pn, ic, mb, it, row["quantity"], row["user_name"]], 1):
            _set_data(ws, ri, ci, val)
    _auto_fit(ws)
    ws.freeze_panes = "A2"


# ═══════════════════  Sheet 5: 用户统计 — 模板格式（4级：用户名→实验室→代号→仪器） ═══════════════════
def _build_user_stats(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "9C27B0"
    if start_s: ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else: ref = datetime.now()
    d_start, d_end = _month_bounds(ref)

    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT wr.user_name, pg.name AS group_name, p.name AS project_name, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY wr.user_name, p.id ORDER BY wr.user_name, pg.sort_order, p.sort_order", rp).fetchall()

    # 4级结构: user → lab → code → {lc:[], gc:[]}
    user_order, user_data = [], {}
    for row in rows:
        un, gn, pn = row["user_name"], row["group_name"], row["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        if un not in user_data: user_data[un] = ([], {}); user_order.append(un)
        if gn not in user_data[un][1]: user_data[un][1][gn] = ([], {}); user_data[un][0].append(gn)
        if pc_code not in user_data[un][1][gn][1]:
            user_data[un][1][gn][1][pc_code] = {'lc': [], 'gc': []}
            user_data[un][1][gn][0].append(pc_code)
        if it == "气相": user_data[un][1][gn][1][pc_code]['gc'].append((ic, fn, row["qty"] or 0))
        else: user_data[un][1][gn][1][pc_code]['lc'].append((ic, fn, row["qty"] or 0))

    # 表头: B=用户名 C=使用实验室 D=项目代号 E=检测方法 F=月检测数量 G=液相 H=气相 I=总量
    CB, CC, CD, CE = _B, _C, _D, _E
    CF, CG, CH, CI = _F, _G, _H, _I
    for col, lb in [(CB, "用户名"), (CC, "使用实验室"), (CD, "项目代号"), (CE, "检测方法"),
                     (CF, "月检测数量"), (CG, "液相检测量"), (CH, "气相检测量"), (CI, "项目检测总量")]:
        _set_header(ws, _HR, col, lb)

    r = _HR + 1
    for un in user_order:
        user_start = r
        for gn in user_data[un][0]:
            lab_start = r
            for pc_code in user_data[un][1][gn][0]:
                inst = user_data[un][1][gn][1][pc_code]
                code_start = r
                for ic, ml, qty in inst['lc'] + inst['gc']:
                    _set_data(ws, r, CB, un)
                    _set_data(ws, r, CC, gn)
                    _set_data(ws, r, CD, pc_code)
                    _set_data(ws, r, CE, ml)
                    _set_data(ws, r, CF, qty if qty else None)
                    r += 1
                code_end = r - 1
                if code_end > code_start:
                    ws.merge_cells(start_row=code_start, start_column=CD, end_row=code_end, end_column=CD)
                    for col in [CG, CH, CI]:
                        ws.merge_cells(start_row=code_start, start_column=col, end_row=code_end, end_column=col)
                lc_end = code_start + len(inst['lc']) - 1 if inst['lc'] else code_start - 1
                gc_start = code_start + len(inst['lc'])
                gc_end = code_end if inst['gc'] else gc_start - 1
                if inst['lc'] and inst['gc']:
                    ws.cell(row=code_start, column=CG).value = f"=SUM({_cl(CF)}{code_start}:{_cl(CF)}{lc_end})"
                    ws.cell(row=code_start, column=CH).value = f"=SUM({_cl(CF)}{gc_start}:{_cl(CF)}{gc_end})"
                elif inst['lc']:
                    ws.cell(row=code_start, column=CG).value = f"=SUM({_cl(CF)}{code_start}:{_cl(CF)}{lc_end})"
                elif inst['gc']:
                    ws.cell(row=code_start, column=CH).value = f"=SUM({_cl(CF)}{gc_start}:{_cl(CF)}{gc_end})"
                ws.cell(row=code_start, column=CI).value = f"=SUM({_cl(CG)}{code_start}:{_cl(CH)}{code_end})"
                for col in [CG, CH, CI]:
                    for br in range(code_start, code_end + 1):
                        ws.cell(row=br, column=col).font = FONT_D
                        ws.cell(row=br, column=col).alignment = ALIGN_C
                        ws.cell(row=br, column=col).border = BORDER_T
            lab_end = r - 1
            if lab_end > lab_start:
                ws.merge_cells(start_row=lab_start, start_column=CC, end_row=lab_end, end_column=CC)
        user_end = r - 1
        if user_end > user_start:
            ws.merge_cells(start_row=user_start, start_column=CB, end_row=user_end, end_column=CB)

    # 总计行
    total_row = r
    _set_data(ws, total_row, CB, "总计")
    _set_data(ws, total_row, CC, "")
    _set_data(ws, total_row, CD, "")
    _set_data(ws, total_row, CE, "")
    sr = f"{_HR + 1}:{total_row - 1}"
    for col in [CF, CG, CH, CI]:
        ws.cell(row=total_row, column=col, value=f"=SUM({_cl(col)}{sr})")
        ws.cell(row=total_row, column=col).font = FONT_D
        ws.cell(row=total_row, column=col).alignment = ALIGN_C
        ws.cell(row=total_row, column=col).border = BORDER_T

    ws.column_dimensions['A'].width = 3
    _auto_fit(ws, min_col=_B, max_col=_I, data_start=_HR)
    for cl, w in WIDTHS.items(): ws.column_dimensions[cl].width = w
    ws.freeze_panes = f"{_cl(CE)}{_HR + 1}"


# ═══════════════════  导出入口 ═══════════════════
@router.get("/excel")
def export_excel(start: str = Query(None), end: str = Query(None), group_id: int = Query(None), db=Depends(get_db)):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "月-汇总"; ws1.sheet_properties.tabColor = "1976D2"
    month_label = _build_monthly_summary(ws1, db, start, end, group_id)
    ws2 = wb.create_sheet(title="每日工作量"); _build_daily_work(ws2, db, start, end, group_id)
    ws3 = wb.create_sheet(title="每周工作量"); _build_weekly_work(ws3, db, start, end, group_id)
    ws4 = wb.create_sheet(title="原始记录"); _build_raw_records(ws4, db, start, end, group_id)
    ws5 = wb.create_sheet(title="用户统计"); _build_user_stats(ws5, db, start, end, group_id)
    output = BytesIO(); wb.save(output); output.seek(0)
    filename = quote(f"工作量统计_{month_label}.xlsx")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
