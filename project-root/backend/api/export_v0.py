"""Excel export — 仿宋16号 月汇总+每周+每日+原始记录+用户"""
import re
from collections import defaultdict
from io import BytesIO
from urllib.parse import quote
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db, get_method_full_name

router = APIRouter(prefix="/export", tags=["export"])

FONT_H = Font(name="仿宋", size=16, bold=True)
FONT_D = Font(name="仿宋", size=16)
FONT_B = Font(name="仿宋", size=16, bold=True)
ALIGN_C = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin")
MEDIUM = Side(style="medium")
BORDER_T = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _cl(n): return get_column_letter(n)

def _auto_fit(ws, min_col=1, max_col=None, data_start=1, tight_cols=None):
    """Auto-fit column widths. tight_cols = set of col indices that only hold short numbers/dates."""
    if max_col is None:
        max_col = ws.max_column
    if tight_cols is None:
        tight_cols = set()
    for col in range(min_col, max_col + 1):
        best = 0
        for row in ws.iter_rows(min_row=data_start, min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    text = str(cell.value)
                    w = sum(2 if '\u4e00' <= c <= '\u9fff' or '\u3000' <= c <= '\u303f'
                            or '\uff00' <= c <= '\uffef' else 1 for c in text)
                    if w > best: best = w
        if col in tight_cols:
            ws.column_dimensions[_cl(col)].width = max(int(best * 1.2) + 2, 5) if best else 6
        else:
            ws.column_dimensions[_cl(col)].width = max(int(best * 1.8) + 6, 10) if best else 10

def _parse_instrument(project_name: str) -> tuple[str, str, str]:
    m = re.match(r'^(.+?)-((?:LC|GC)-\d+)(.*)$', project_name)
    if m:
        prefix, code, suffix = m.group(1), m.group(2), m.group(3)
        itype = "液相" if code.upper().startswith("LC") else "气相"
        return (f"{prefix}-{suffix}" if suffix else prefix, code, itype)
    return (project_name, "", "其他")

def _extract_project_code(n): return n.split("-", 1)[0] if "-" in n else n

def _week_ranges(start: datetime, end: datetime) -> list[tuple[str, datetime, datetime]]:
    weeks = []
    cur = start - timedelta(days=start.weekday())
    while cur <= end:
        wk_end = cur + timedelta(days=6)
        label = f"{cur.strftime('%m.%d')}-{wk_end.strftime('%m.%d')}"
        weeks.append((label, cur, wk_end))
        cur = wk_end + timedelta(days=1)
    return weeks

def _month_bounds(ref_date):
    d_start = ref_date.replace(day=1)
    if d_start.month == 12:
        d_end = d_start.replace(year=d_start.year+1, month=1, day=1) - timedelta(days=1)
    else:
        d_end = d_start.replace(month=d_start.month+1, day=1) - timedelta(days=1)
    return d_start, d_end

# ═══════════════════  Sheet 1: 月-汇总 ═══════════════════
def _build_monthly_summary(ws, db, start_s, end_s, group_id):
    if start_s:
        ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else:
        ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    month_label = f"{d_start.year}年{d_start.month}月"

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None:
        pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT p.id AS project_id, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY p.id", rp).fetchall()
    proj_monthly = {r["project_id"]: (r["qty"] or 0) for r in records}

    lab_order, lab_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        qty = proj_monthly.get(proj["id"], 0)
        if gn not in lab_data:
            lab_data[gn] = {}; lab_order.append(gn)
        if pc_code not in lab_data[gn]:
            lab_data[gn][pc_code] = {'lc': [], 'gc': []}
        if it == "气相":
            lab_data[gn][pc_code]['gc'].append((ic, fn, pn, qty))
        else:
            lab_data[gn][pc_code]['lc'].append((ic, fn, pn, qty))

    CB, CC, CD, CE = 2, 3, 4, 5
    CF = 6; CG = 7; CH = 8; CI = 9
    CJ = 10; CK = 11; CL = 12; CM = 13
    LAST = CM; HR = 2

    hd = {CB: "使用实验室", CC: "项目代号", CD: "液相仪器", CE: "检测方法",
          CF: "月检测数量", CG: "液相检测量", CH: "气相检测量", CI: "项目检测总量",
          CJ: None, CK: "液相检测总量", CL: "气相检测总量", CM: "检测总量"}
    for ci, lb in hd.items():
        if lb is None: continue
        c = ws.cell(row=HR, column=ci, value=lb)
        c.font = FONT_B if ci in (CK, CL, CM) else FONT_H
        c.alignment = ALIGN_C; c.border = BORDER_T

    r = HR + 1
    for lab_name in lab_order:
        pd = lab_data[lab_name]; lsr = r
        for pc_code, inst in pd.items():
            psr = r
            for ic, ml, pn, qty in inst['lc'] + inst['gc']:
                for col, val in [(CB, lab_name), (CC, pc_code), (CD, ic), (CE, ml),
                                 (CF, qty if qty else None)]:
                    c = ws.cell(row=r, column=col, value=val)
                    c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                r += 1
            per = r - 1
            if psr > per: continue
            if per > psr:
                ws.merge_cells(start_row=psr, start_column=CC, end_row=per, end_column=CC)
            for col in [CG, CH, CI]:
                if per > psr:
                    ws.merge_cells(start_row=psr, start_column=col, end_row=per, end_column=col)

            lc_range = f"{_cl(CF)}{psr}:{_cl(CF)}{psr + len(inst['lc']) - 1}" if inst['lc'] else None
            gc_range = f"{_cl(CF)}{psr + len(inst['lc'])}:{_cl(CF)}{per}" if inst['gc'] else None
            if lc_range and gc_range:
                ws.cell(row=psr, column=CG).value = f"=SUM({lc_range})"
                ws.cell(row=psr, column=CH).value = f"=SUM({gc_range})"
            elif lc_range:
                ws.cell(row=psr, column=CG).value = f"=SUM({lc_range})"
            elif gc_range:
                ws.cell(row=psr, column=CH).value = f"=SUM({gc_range})"
            ws.cell(row=psr, column=CI).value = f"=SUM({_cl(CG)}{psr}:{_cl(CH)}{per})"

            for col in [CG, CH, CI]:
                for br in range(psr, per + 1):
                    ws.cell(row=br, column=col).font = FONT_D
                    ws.cell(row=br, column=col).alignment = ALIGN_C
                    ws.cell(row=br, column=col).border = BORDER_T

        ler = r - 1
        if ler > lsr:
            ws.merge_cells(start_row=lsr, start_column=CB, end_row=ler, end_column=CB)

    total_row = r
    for col, label in [(CB, "总计"), (CC, ""), (CD, ""), (CE, ""),
                        (CF, f"=SUM({_cl(CF)}{HR + 1}:{_cl(CF)}{total_row - 1})"),
                        (CG, f"=SUM({_cl(CG)}{HR + 1}:{_cl(CG)}{total_row - 1})"),
                        (CH, f"=SUM({_cl(CH)}{HR + 1}:{_cl(CH)}{total_row - 1})"),
                        (CI, f"=SUM({_cl(CI)}{HR + 1}:{_cl(CI)}{total_row - 1})")]:
        c = ws.cell(row=total_row, column=col, value=label)
        c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T

    for col, ref_col in [(CK, CG), (CL, CH), (CM, CI)]:
        c = ws.cell(row=total_row, column=col, value=f"={_cl(ref_col)}{total_row}")
        c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T

    _auto_fit(ws, min_col=2, max_col=LAST, data_start=2)
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['J'].width = 3
    ws.freeze_panes = f"{_cl(CE)}{HR + 1}"
    return month_label


# ═══════════════════  Sheet 2: 每日工作量 ═══════════════════
def _build_daily_work(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "1976D2"
    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT date(wr.recorded_at) AS work_day, pg.name AS group_name, p.name AS project_name, SUM(wr.quantity) AS qty, GROUP_CONCAT(DISTINCT wr.user_name) AS users FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY date(wr.recorded_at), pg.id, p.id ORDER BY work_day, pg.sort_order, p.sort_order", rp).fetchall()

    hd = ["日期", "实验室", "项目代码", "仪器", "检测方法", "类型", "数量", "录入人"]
    for ci, lb in enumerate(hd, 1):
        c = ws.cell(row=1, column=ci, value=lb); c.font = FONT_H; c.alignment = ALIGN_C; c.border = BORDER_T

    for ri, row in enumerate(rows, 2):
        gn, pn = row["group_name"], row["project_name"]
        pc = _extract_project_code(pn); mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        wd = row["work_day"]; wdd = wd if isinstance(wd, str) else str(wd)
        vals = [wdd, gn, pc, ic, fn, it, row["qty"], row["users"] or ""]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val); c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T

    _auto_fit(ws)
    ws.freeze_panes = "A2"


# ═══════════════════  Sheet 3: 每周工作量 (same format as 月-汇总) ═══════════════════
def _build_weekly_work(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "43A047"
    if start_s:
        ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else:
        ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    weeks = _week_ranges(d_start, d_end)
    num_weeks = len(weeks)

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None:
        pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT p.id AS project_id, date(wr.recorded_at) AS work_day, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY p.id, date(wr.recorded_at)", rp).fetchall()

    proj_weekly = defaultdict(lambda: defaultdict(int))
    for row in records:
        rd = datetime.strptime(row["work_day"], "%Y-%m-%d")
        for wi, (_, ws_d, we_d) in enumerate(weeks):
            if ws_d <= rd <= we_d:
                proj_weekly[row["project_id"]][wi] += (row["qty"] or 0)
                break

    lab_order, lab_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        dm = proj_weekly.get(proj["id"], {})
        if gn not in lab_data:
            lab_data[gn] = {}; lab_order.append(gn)
        if pc_code not in lab_data[gn]:
            lab_data[gn][pc_code] = {'lc': [], 'gc': []}
        if it == "气相":
            lab_data[gn][pc_code]['gc'].append((ic, fn, pn, dict(dm)))
        else:
            lab_data[gn][pc_code]['lc'].append((ic, fn, pn, dict(dm)))

    CB, CC, CD, CE = 2, 3, 4, 5
    CW = 6; CMT = CW + num_weeks
    CGC = CMT + 1; CLR = CGC + 1; CGR = CLR + 1; CGD = CGR + 1
    LAST = CGD; HR = 2
    tight = set(range(CW, CW + num_weeks))

    hd = {CB: "使用实验室", CC: "项目代号", CD: "液相仪器", CE: "检测方法",
          CMT: "月检测数量", CGC: "气相检测量", CLR: "液相检测总量", CGR: "气相检测总量", CGD: "检测总量"}
    for ci, lb in hd.items():
        c = ws.cell(row=HR, column=ci, value=lb)
        c.font = FONT_B if ci in (CLR, CGR, CGD) else FONT_H
        c.alignment = ALIGN_C; c.border = BORDER_T

    if num_weeks > 0:
        ws.merge_cells(start_row=HR, start_column=CW, end_row=HR, end_column=CW + num_weeks - 1)
        c = ws.cell(row=HR, column=CW, value=f"周汇总（{d_start.month}月）")
        c.font = FONT_H; c.alignment = ALIGN_C
        for dc in range(CW, CW + num_weeks): ws.cell(row=HR, column=dc).border = BORDER_T

    R3 = HR + 1
    for i, (label, _, _) in enumerate(weeks):
        c = ws.cell(row=R3, column=CW + i, value=label); c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
    for ci in [CB, CC, CD, CE, CMT, CGC, CLR, CGR, CGD]:
        ws.merge_cells(start_row=HR, start_column=ci, end_row=R3, end_column=ci)
        ws.cell(row=R3, column=ci).border = BORDER_T

    DS = R3 + 1; r = DS; srows = []; rlc, rgc = 0, 0
    for lab_name in lab_order:
        pd = lab_data[lab_name]; lsr = r
        for pc_code, inst in pd.items():
            psr = r; plc, pgc = 0, 0
            for ic, ml, pn, dm in inst['lc'] + inst['gc']:
                is_gc = pn in {x[2] for x in inst['gc']}
                rt = 0
                for col, val in [(CB, lab_name), (CC, pc_code), (CD, ic), (CE, ml)]:
                    c = ws.cell(row=r, column=col, value=val)
                    c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                for wi in range(num_weeks):
                    qty = dm.get(wi, 0); rt += qty
                    c = ws.cell(row=r, column=CW + wi, value=qty if qty else None)
                    c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                c = ws.cell(row=r, column=CMT, value=rt if rt else None)
                c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                if is_gc: pgc += rt
                else: plc += rt
                r += 1
            per = r - 1
            if psr > per: continue
            if per > psr: ws.merge_cells(start_row=psr, start_column=CC, end_row=per, end_column=CC)
            rlc += plc; rgc += pgc

        ler = r - 1
        if ler > lsr: ws.merge_cells(start_row=lsr, start_column=CB, end_row=ler, end_column=CB)
        ws.merge_cells(start_row=r, start_column=CB, end_row=r, end_column=CC)
        c = ws.cell(row=r, column=CB, value=f"{lab_name} 小计"); c.font = FONT_B; c.alignment = ALIGN_C
        for cc in range(CB, LAST + 1):
            ws.cell(row=r, column=cc).font = FONT_B
            ws.cell(row=r, column=cc).alignment = ALIGN_C
            ws.cell(row=r, column=cc).border = BORDER_T
        for wi in range(num_weeks):
            cl = _cl(CW + wi)
            ws.cell(row=r, column=CW + wi).value = f"=SUM({cl}{lsr}:{cl}{ler})"
        mtc = _cl(CMT); ws.cell(row=r, column=CMT).value = f"=SUM({mtc}{lsr}:{mtc}{ler})"
        ws.cell(row=r, column=CLR).value = rlc if rlc else None
        ws.cell(row=r, column=CGR).value = rgc if rgc else None
        ws.cell(row=r, column=CGD).value = (rlc + rgc) if (rlc + rgc) else None
        for cc in range(CB, LAST + 1):
            b = ws.cell(row=r, column=cc).border
            ws.cell(row=r, column=cc).border = Border(left=b.left, right=b.right, top=b.top, bottom=MEDIUM)
        srows.append(r); r += 1

    ws.merge_cells(start_row=r, start_column=CB, end_row=r, end_column=CC)
    c = ws.cell(row=r, column=CB, value="总计"); c.font = FONT_B; c.alignment = ALIGN_C
    for cc in range(CB, LAST + 1):
        ws.cell(row=r, column=cc).font = FONT_B
        ws.cell(row=r, column=cc).alignment = ALIGN_C
        ws.cell(row=r, column=cc).border = BORDER_T
    if srows:
        for wi in range(num_weeks):
            cl = _cl(CW + wi)
            ws.cell(row=r, column=CW + wi).value = "=SUM(" + ",".join(f"{cl}{sr}" for sr in srows) + ")"
        mtc = _cl(CMT); ws.cell(row=r, column=CMT).value = "=SUM(" + ",".join(f"{mtc}{sr}" for sr in srows) + ")"
    ws.cell(row=r, column=CLR).value = rlc if rlc else None
    ws.cell(row=r, column=CGR).value = rgc if rgc else None
    ws.cell(row=r, column=CGD).value = (rlc + rgc) if (rlc + rgc) else None
    for cc in range(CB, LAST + 1):
        b = ws.cell(row=r, column=cc).border
        ws.cell(row=r, column=cc).border = Border(left=b.left, right=b.right, top=b.top, bottom=MEDIUM)

    _auto_fit(ws, min_col=2, max_col=LAST, data_start=2, tight_cols=tight)
    ws.column_dimensions['A'].width = 3
    ws.freeze_panes = f"{_cl(CE)}{DS}"


# ═══════════════════  Sheet 4: 原始记录 ═══════════════════
def _build_raw_records(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "FF9800"
    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT pg.name AS group_name, p.name AS project_name, wr.user_name, wr.quantity, wr.recorded_at FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} ORDER BY wr.recorded_at, pg.sort_order, p.sort_order", rp).fetchall()

    hd = ["序号", "日期", "实验室", "项目名称", "仪器", "检测方法", "仪器类型", "数量", "录入人"]
    for ci, lb in enumerate(hd, 1):
        c = ws.cell(row=1, column=ci, value=lb); c.font = FONT_H; c.alignment = ALIGN_C; c.border = BORDER_T

    for ri, row in enumerate(rows, 2):
        gn, pn = row["group_name"], row["project_name"]
        mb, ic, it = _parse_instrument(pn)
        rd = row["recorded_at"]; rdd = rd[:10] if isinstance(rd, str) else str(rd)[:10]
        vals = [ri - 1, rdd, gn, pn, ic, mb, it, row["quantity"], row["user_name"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val); c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T

    _auto_fit(ws)
    ws.freeze_panes = "A2"


# ═══════════════════  Sheet 5: 用户统计 (same as 月-汇总 + weekly) ═══════════════════
def _build_user_stats(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "9C27B0"
    if start_s:
        ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else:
        ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    weeks = _week_ranges(d_start, d_end)
    num_weeks = len(weeks)

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None:
        pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    # Per-user per-project daily records
    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT wr.user_name, p.id AS project_id, date(wr.recorded_at) AS work_day, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY wr.user_name, p.id, date(wr.recorded_at)", rp).fetchall()

    # user_name -> project_id -> {week_index: qty}
    user_proj_weekly = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for row in records:
        un, pid = row["user_name"], row["project_id"]
        rd = datetime.strptime(row["work_day"], "%Y-%m-%d")
        for wi, (_, ws_d, we_d) in enumerate(weeks):
            if ws_d <= rd <= we_d:
                user_proj_weekly[un][pid][wi] += (row["qty"] or 0)
                break

    # Organize: user_name -> (lab_order, {lab: (proj_order, {pc: {lc/gc})})})
    user_order, user_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        # Get per-user weekly data for this project
        user_project_wk = user_proj_weekly.get(proj["id"], {})  # not used directly, accessed via user
        # We need to iterate users. But user_proj_weekly is indexed by (user, project).
        # Instead, iterate all users who have this project:
        for un in user_proj_weekly:
            dm = user_proj_weekly[un].get(proj["id"], {})
            if un not in user_data:
                user_data[un] = ([], {})
                user_order.append(un)
            if gn not in user_data[un][1]:
                user_data[un][1][gn] = ([], {})
                user_data[un][0].append(gn)
            if pc_code not in user_data[un][1][gn][1]:
                user_data[un][1][gn][1][pc_code] = {'lc': [], 'gc': []}
                user_data[un][1][gn][0].append(pc_code)
            if it == "气相":
                user_data[un][1][gn][1][pc_code]['gc'].append((ic, fn, dict(dm)))
            else:
                user_data[un][1][gn][1][pc_code]['lc'].append((ic, fn, dict(dm)))
            break  # break the user loop since we only process one project at a time
    # Wait, this logic is wrong. We need to restructure.

    # Better approach: rebuild from all_projects but check each user's data
    user_order2, user_data2 = [], {}
    all_users = set()
    for row in records:
        all_users.add(row["user_name"])
    # Also check for users with no records but tracked
    all_users_sorted = sorted(all_users)

    for un in all_users_sorted:
        user_data2[un] = ([], {})
        user_order2.append(un)
        for proj in all_projects:
            gn, pn = proj["group_name"], proj["project_name"]
            pc_code = _extract_project_code(pn)
            mb, ic, it = _parse_instrument(pn)
            fn = get_method_full_name(gn, pn) or mb
            dm = user_proj_weekly.get(un, {}).get(proj["id"], {})
            if gn not in user_data2[un][1]:
                user_data2[un][1][gn] = ([], {})
                user_data2[un][0].append(gn)
            if pc_code not in user_data2[un][1][gn][1]:
                user_data2[un][1][gn][1][pc_code] = {'lc': [], 'gc': []}
                user_data2[un][1][gn][0].append(pc_code)
            if it == "气相":
                user_data2[un][1][gn][1][pc_code]['gc'].append((ic, fn, dict(dm)))
            else:
                user_data2[un][1][gn][1][pc_code]['lc'].append((ic, fn, dict(dm)))

    user_order, user_data = user_order2, user_data2

    # Columns: B=用户名 C=实验室 D=项目代号 E=仪器 F=检测方法 G..=每周 ...
    CB, CC, CD, CE, CF = 2, 3, 4, 5, 6
    CW = 7; CMT = CW + num_weeks
    CGC = CMT + 1; CLR = CGC + 1; CGR = CLR + 1; CGD = CGR + 1
    LAST = CGD; HR = 2
    tight = set(range(CW, CW + num_weeks))

    hd = {CB: "用户名", CC: "使用实验室", CD: "项目代号", CE: "液相仪器", CF: "检测方法",
          CMT: "月检测数量", CGC: "气相检测量", CLR: "液相检测总量", CGR: "气相检测总量", CGD: "检测总量"}
    for ci, lb in hd.items():
        c = ws.cell(row=HR, column=ci, value=lb)
        c.font = FONT_B if ci in (CLR, CGR, CGD) else FONT_H
        c.alignment = ALIGN_C; c.border = BORDER_T

    if num_weeks > 0:
        ws.merge_cells(start_row=HR, start_column=CW, end_row=HR, end_column=CW + num_weeks - 1)
        c = ws.cell(row=HR, column=CW, value=f"周汇总（{d_start.month}月）")
        c.font = FONT_H; c.alignment = ALIGN_C
        for dc in range(CW, CW + num_weeks): ws.cell(row=HR, column=dc).border = BORDER_T

    R3 = HR + 1
    for i, (label, _, _) in enumerate(weeks):
        c = ws.cell(row=R3, column=CW + i, value=label); c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
    for ci in [CB, CC, CD, CE, CF, CMT, CGC, CLR, CGR, CGD]:
        ws.merge_cells(start_row=HR, start_column=ci, end_row=R3, end_column=ci)
        ws.cell(row=R3, column=ci).border = BORDER_T

    DS = R3 + 1; r = DS
    for un in user_order:
        lab_order, lab_map = user_data[un]
        user_start = r
        for gn in lab_order:
            proj_order, proj_dict = lab_map[gn]
            lab_start = r
            for pc in proj_order:
                inst = proj_dict[pc]
                proj_start = r
                for ic, ml, dm in inst['lc'] + inst['gc']:
                    rt = 0
                    for col, val in [(CB, un), (CC, gn), (CD, pc), (CE, ic), (CF, ml)]:
                        c = ws.cell(row=r, column=col, value=val)
                        c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                    for wi in range(num_weeks):
                        qty = dm.get(wi, 0); rt += qty
                        c = ws.cell(row=r, column=CW + wi, value=qty if qty else None)
                        c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                    c = ws.cell(row=r, column=CMT, value=rt if rt else None)
                    c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                    r += 1
                proj_end = r - 1
                if proj_end > proj_start:
                    ws.merge_cells(start_row=proj_start, start_column=CD, end_row=proj_end, end_column=CD)
            lab_end = r - 1
            if lab_end > lab_start:
                ws.merge_cells(start_row=lab_start, start_column=CC, end_row=lab_end, end_column=CC)
        user_end = r - 1
        if user_end > user_start:
            ws.merge_cells(start_row=user_start, start_column=CB, end_row=user_end, end_column=CB)

    _auto_fit(ws, min_col=2, max_col=LAST, data_start=2, tight_cols=tight)
    ws.column_dimensions['A'].width = 3
    ws.freeze_panes = f"{_cl(CF)}{DS}"


@router.get("/excel")
def export_excel(start: str = Query(None), end: str = Query(None), group_id: int = Query(None), db=Depends(get_db)):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "月-汇总"; ws1.sheet_properties.tabColor = "1976D2"
    month_label = _build_monthly_summary(ws1, db, start, end, group_id)
    ws2 = wb.create_sheet(title="每日工作量"); _build_daily_work(ws2, db, start, end, group_id)
    ws3 = wb.create_sheet(title="每周工作量"); _build_weekly_work(ws3, db, start, end, group_id)
    ws4 = wb.create_sheet(title="原始记录"); _build_raw_records(ws4, db, start, end, group_id)
    ws5 = wb.create_sheet(title="用户统计"); _build_user_stats(ws5, db, start, end, group_id)
    output = BytesIO(); wb.save(output); output.seek(0)
    filename = quote(f"工作量统计_{month_label}.xlsx")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
