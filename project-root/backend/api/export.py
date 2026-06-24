"""Excel export — 仿宋16号 严格匹配汇总模板 + 其他Sheet"""
import re
from collections import defaultdict
from io import BytesIO
from urllib.parse import quote
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db, get_method_full_name

router = APIRouter(prefix="/export", tags=["export"])

FONT_H = Font(name="仿宋", size=16, bold=True)
FONT_D = Font(name="仿宋", size=16)
FONT_R = Font(name="仿宋", size=16, bold=False)
FONT_B = Font(name="仿宋", size=16, bold=True)
ALIGN_C = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin")
BORDER_T = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Exact column widths from template (A-J only, K/L/M are unused spacers)
WIDTHS = {'A': 8.89, 'B': 24.89, 'C': 18.0, 'D': 17.44, 'E': 43.66,
          'F': 18.66, 'I': 22.11, 'J': 16.78}

def _cl(n): return get_column_letter(n)

def _parse_instrument(project_name: str) -> tuple[str, str, str]:
    m = re.match(r'^(.+?)-((?:LC|GC)-\d+)(.*)$', project_name)
    if m:
        prefix, code, suffix = m.group(1), m.group(2), m.group(3)
        itype = "液相" if code.upper().startswith("LC") else "气相"
        return (f"{prefix}-{suffix}" if suffix else prefix, code, itype)
    return (project_name, "", "其他")

def _extract_project_code(n): return n.split("-", 1)[0] if "-" in n else n

def _week_ranges(start: datetime, end: datetime) -> list[tuple[str, datetime, datetime]]:
    weeks = []
    cur = start - timedelta(days=start.weekday())
    while cur <= end:
        wk_end = cur + timedelta(days=6)
        label = f"{cur.strftime('%m.%d')}-{wk_end.strftime('%m.%d')}"
        weeks.append((label, cur, wk_end))
        cur = wk_end + timedelta(days=1)
    return weeks

def _month_bounds(ref_date):
    d_start = ref_date.replace(day=1)
    if d_start.month == 12:
        d_end = d_start.replace(year=d_start.year+1, month=1, day=1) - timedelta(days=1)
    else:
        d_end = d_start.replace(month=d_start.month+1, day=1) - timedelta(days=1)
    return d_start, d_end

# ═══════════════════  Sheet 1: 月-汇总 (EXACT template match) ═══════════════════
def _build_monthly_summary(ws, db, start_s, end_s, group_id):
    if start_s:
        ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else:
        ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    month_label = f"{d_start.year}年{d_start.month}月"

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None:
        pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT p.id AS project_id, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY p.id", rp).fetchall()
    proj_monthly = {r["project_id"]: (r["qty"] or 0) for r in records}

    lab_order, lab_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        qty = proj_monthly.get(proj["id"], 0)
        if gn not in lab_data:
            lab_data[gn] = {}; lab_order.append(gn)
        if pc_code not in lab_data[gn]:
            lab_data[gn][pc_code] = {'lc': [], 'gc': []}
        if it == "气相":
            lab_data[gn][pc_code]['gc'].append((ic, fn, pn, qty))
        else:
            lab_data[gn][pc_code]['lc'].append((ic, fn, pn, qty))

    # Columns: B=使用实验室 C=项目代号 D=液相仪器 E=检测方法
    # F=月检测数量 G=液相检测量 H=气相检测量 I=项目检测总量 J=空
    CB, CC, CD, CE = 2, 3, 4, 5
    CF = 6; CG = 7; CH = 8; CI = 9; HR = 2

    # Headers: B-I bold (matching template)
    for col, label in [(CB, "使用实验室"), (CC, "项目代号"), (CD, "液相仪器"), (CE, "检测方法"),
                        (CF, "月检测数量"), (CG, "液相检测量"), (CH, "气相检测量"), (CI, "项目检测总量")]:
        c = ws.cell(row=HR, column=col, value=label)
        c.font = FONT_H; c.alignment = ALIGN_C; c.border = BORDER_T

    # Data rows — flat, no subtotal rows, only merge cells
    r = HR + 1
    for lab_name in lab_order:
        pd = lab_data[lab_name]
        lab_start = r
        for pc_code, inst in pd.items():
            psr = r
            for ic, ml, pn, qty in inst['lc'] + inst['gc']:
                for col, val in [(CB, lab_name), (CC, pc_code), (CD, ic), (CE, ml),
                                 (CF, qty if qty else None)]:
                    c = ws.cell(row=r, column=col, value=val)
                    c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                r += 1
            per = r - 1
            if psr > per: continue
            if per > psr:
                ws.merge_cells(start_row=psr, start_column=CC, end_row=per, end_column=CC)
            for col in [CG, CH, CI]:
                if per > psr:
                    ws.merge_cells(start_row=psr, start_column=col, end_row=per, end_column=col)

            lc_range = f"{_cl(CF)}{psr}:{_cl(CF)}{psr + len(inst['lc']) - 1}" if inst['lc'] else None
            gc_range = f"{_cl(CF)}{psr + len(inst['lc'])}:{_cl(CF)}{per}" if inst['gc'] else None
            if lc_range and gc_range:
                ws.cell(row=psr, column=CG).value = f"=SUM({lc_range})"
                ws.cell(row=psr, column=CH).value = f"=SUM({gc_range})"
            elif lc_range:
                ws.cell(row=psr, column=CG).value = f"=SUM({lc_range})"
            elif gc_range:
                ws.cell(row=psr, column=CH).value = f"=SUM({gc_range})"
            ws.cell(row=psr, column=CI).value = f"=SUM({_cl(CG)}{psr}:{_cl(CH)}{per})"

            for col in [CG, CH, CI]:
                for br in range(psr, per + 1):
                    ws.cell(row=br, column=col).font = FONT_D
                    ws.cell(row=br, column=col).alignment = ALIGN_C
                    ws.cell(row=br, column=col).border = BORDER_T

        ler = r - 1
        if ler > lab_start:
            ws.merge_cells(start_row=lab_start, start_column=CB, end_row=ler, end_column=CB)

    total_row = r

    # Grand total row — 总计 is regular (matching template)
    for col, label in [(CB, "总计"), (CC, ""), (CD, ""), (CE, ""),
                        (CF, f"=SUM({_cl(CF)}{HR + 1}:{_cl(CF)}{total_row - 1})"),
                        (CG, f"=SUM({_cl(CG)}{HR + 1}:{_cl(CG)}{total_row - 1})"),
                        (CH, f"=SUM({_cl(CH)}{HR + 1}:{_cl(CH)}{total_row - 1})"),
                        (CI, f"=SUM({_cl(CI)}{HR + 1}:{_cl(CI)}{total_row - 1})")]:
        c = ws.cell(row=total_row, column=col, value=label)
        c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T

    # Apply exact template column widths
    for col_letter, w in WIDTHS.items():
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = f"{_cl(CE)}{HR + 1}"
    return month_label


# ═══════════════════  Sheet 2: 每日工作量 ═══════════════════
def _build_daily_work(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "1976D2"
    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT date(wr.recorded_at) AS work_day, pg.name AS group_name, p.name AS project_name, SUM(wr.quantity) AS qty, GROUP_CONCAT(DISTINCT wr.user_name) AS users FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY date(wr.recorded_at), pg.id, p.id ORDER BY work_day, pg.sort_order, p.sort_order", rp).fetchall()

    hd = ["日期", "实验室", "项目代码", "仪器", "检测方法", "类型", "数量", "录入人"]
    for ci, lb in enumerate(hd, 1):
        c = ws.cell(row=1, column=ci, value=lb); c.font = FONT_H; c.alignment = ALIGN_C; c.border = BORDER_T

    for ri, row in enumerate(rows, 2):
        gn, pn = row["group_name"], row["project_name"]
        pc = _extract_project_code(pn); mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        wd = row["work_day"]; wdd = wd if isinstance(wd, str) else str(wd)
        vals = [wdd, gn, pc, ic, fn, it, row["qty"], row["users"] or ""]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val); c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T

    for col, w in [(1, 12), (2, 14), (3, 12), (4, 10), (5, 43), (6, 8), (7, 8), (8, 12)]:
        ws.column_dimensions[_cl(col)].width = w
    ws.freeze_panes = "A2"


# ═══════════════════  Sheet 3: 每周工作量 ═══════════════════
def _build_weekly_work(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "43A047"
    if start_s:
        ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else:
        ref = datetime.now()
    d_start, d_end = _month_bounds(ref)
    weeks = _week_ranges(d_start, d_end)
    num_weeks = len(weeks)

    pc = ["p.is_active = 1"]; pp = []
    if group_id is not None:
        pc.append("pg.id = ?"); pp.append(group_id)
    all_projects = db.execute(
        f"SELECT p.id, p.name AS project_name, pg.name AS group_name, pg.sort_order AS gs, p.sort_order AS ps FROM projects p JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(pc)} ORDER BY gs, ps", pp).fetchall()

    rc = ["wr.deleted_at IS NULL"]; rp = []
    rc.append("wr.recorded_at >= ?"); rp.append(d_start.strftime("%Y-%m-%d"))
    rc.append("wr.recorded_at <= ?"); rp.append(d_end.strftime("%Y-%m-%d") + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    records = db.execute(
        f"SELECT p.id AS project_id, date(wr.recorded_at) AS work_day, SUM(wr.quantity) AS qty FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY p.id, date(wr.recorded_at)", rp).fetchall()

    proj_weekly = defaultdict(lambda: defaultdict(int))
    for row in records:
        rd = datetime.strptime(row["work_day"], "%Y-%m-%d")
        for wi, (_, ws_d, we_d) in enumerate(weeks):
            if ws_d <= rd <= we_d:
                proj_weekly[row["project_id"]][wi] += (row["qty"] or 0)
                break

    lab_order, lab_data = [], {}
    for proj in all_projects:
        gn, pn = proj["group_name"], proj["project_name"]
        pc_code = _extract_project_code(pn)
        mb, ic, it = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        dm = proj_weekly.get(proj["id"], {})
        if gn not in lab_data: lab_data[gn] = {}; lab_order.append(gn)
        if pc_code not in lab_data[gn]: lab_data[gn][pc_code] = {'lc': [], 'gc': []}
        if it == "气相": lab_data[gn][pc_code]['gc'].append((ic, fn, pn, dict(dm)))
        else: lab_data[gn][pc_code]['lc'].append((ic, fn, pn, dict(dm)))

    CB, CC, CD, CE = 2, 3, 4, 5
    CW = 6; CMT = CW + num_weeks
    CGC = CMT + 1; CLR = CGC + 1; CGR = CLR + 1; CGD = CGR + 1
    LAST = CGD; HR = 2

    for col, label in [(CB, "使用实验室"), (CC, "项目代号"), (CD, "液相仪器"), (CE, "检测方法"),
                        (CMT, "月检测数量"), (CGC, "气相检测量")]:
        c = ws.cell(row=HR, column=col, value=label)
        c.font = FONT_H; c.alignment = ALIGN_C; c.border = BORDER_T
    for col, label in [(CLR, "液相检测总量"), (CGR, "气相检测总量"), (CGD, "检测总量")]:
        c = ws.cell(row=HR, column=col, value=label)
        c.font = FONT_R; c.alignment = ALIGN_C; c.border = BORDER_T

    if num_weeks > 0:
        ws.merge_cells(start_row=HR, start_column=CW, end_row=HR, end_column=CW + num_weeks - 1)
        c = ws.cell(row=HR, column=CW, value=f"周汇总（{d_start.month}月）")
        c.font = FONT_H; c.alignment = ALIGN_C
        for dc in range(CW, CW + num_weeks): ws.cell(row=HR, column=dc).border = BORDER_T

    R3 = HR + 1
    for i, (label, _, _) in enumerate(weeks):
        c = ws.cell(row=R3, column=CW + i, value=label); c.font = FONT_R; c.alignment = ALIGN_C; c.border = BORDER_T
    for ci in [CB, CC, CD, CE, CMT, CGC, CLR, CGR, CGD]:
        ws.merge_cells(start_row=HR, start_column=ci, end_row=R3, end_column=ci)
        ws.cell(row=R3, column=ci).border = BORDER_T

    DS = R3 + 1; r = DS; srows = []; rlc, rgc = 0, 0
    for lab_name in lab_order:
        pd = lab_data[lab_name]; lsr = r
        for pc_code, inst in pd.items():
            psr = r; plc, pgc = 0, 0
            for ic, ml, pn, dm in inst['lc'] + inst['gc']:
                is_gc = pn in {x[2] for x in inst['gc']}; rt = 0
                for col, val in [(CB, lab_name), (CC, pc_code), (CD, ic), (CE, ml)]:
                    c = ws.cell(row=r, column=col, value=val)
                    c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                for wi in range(num_weeks):
                    qty = dm.get(wi, 0); rt += qty
                    c = ws.cell(row=r, column=CW + wi, value=qty if qty else None)
                    c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                c = ws.cell(row=r, column=CMT, value=rt if rt else None)
                c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
                if is_gc: pgc += rt
                else: plc += rt
                r += 1
            per = r - 1
            if psr > per: continue
            if per > psr: ws.merge_cells(start_row=psr, start_column=CC, end_row=per, end_column=CC)
            rlc += plc; rgc += pgc
        ler = r - 1
        if ler > lsr: ws.merge_cells(start_row=lsr, start_column=CB, end_row=ler, end_column=CB)
        # Subtotal row
        ws.merge_cells(start_row=r, start_column=CB, end_row=r, end_column=CC)
        c = ws.cell(row=r, column=CB, value=f"{lab_name} 小计"); c.font = FONT_H; c.alignment = ALIGN_C
        for cc in range(CB, LAST + 1):
            ws.cell(row=r, column=cc).font = FONT_H
            ws.cell(row=r, column=cc).alignment = ALIGN_C; ws.cell(row=r, column=cc).border = BORDER_T
        for wi in range(num_weeks):
            cl = _cl(CW + wi); ws.cell(row=r, column=CW + wi).value = f"=SUM({cl}{lsr}:{cl}{ler})"
        mtc = _cl(CMT); ws.cell(row=r, column=CMT).value = f"=SUM({mtc}{lsr}:{mtc}{ler})"
        srows.append(r); r += 1

    # Grand total
    ws.merge_cells(start_row=r, start_column=CB, end_row=r, end_column=CC)
    c = ws.cell(row=r, column=CB, value="总计"); c.font = FONT_B; c.alignment = ALIGN_C
    for cc in range(CB, LAST + 1):
        ws.cell(row=r, column=cc).font = FONT_H
        ws.cell(row=r, column=cc).alignment = ALIGN_C; ws.cell(row=r, column=cc).border = BORDER_T
    if srows:
        for wi in range(num_weeks):
            cl = _cl(CW + wi); ws.cell(row=r, column=CW + wi).value = "=SUM(" + ",".join(f"{cl}{sr}" for sr in srows) + ")"
        mtc = _cl(CMT); ws.cell(row=r, column=CMT).value = "=SUM(" + ",".join(f"{mtc}{sr}" for sr in srows) + ")"

    for col, w in [(1, 3), (2, 24), (3, 16), (4, 10), (5, 43)]:
        ws.column_dimensions[_cl(col)].width = w
    for wi in range(num_weeks):
        ws.column_dimensions[_cl(CW + wi)].width = 8
    ws.freeze_panes = f"{_cl(CE)}{DS}"


# ═══════════════════  Sheet 4: 原始记录 ═══════════════════
def _build_raw_records(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "FF9800"
    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT pg.name AS group_name, p.name AS project_name, wr.user_name, wr.quantity, wr.recorded_at FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} ORDER BY wr.recorded_at, pg.sort_order, p.sort_order", rp).fetchall()

    hd = ["序号", "日期", "实验室", "项目名称", "仪器", "检测方法", "仪器类型", "数量", "录入人"]
    for ci, lb in enumerate(hd, 1):
        c = ws.cell(row=1, column=ci, value=lb); c.font = FONT_H; c.alignment = ALIGN_C; c.border = BORDER_T

    for ri, row in enumerate(rows, 2):
        gn, pn = row["group_name"], row["project_name"]
        mb, ic, it = _parse_instrument(pn)
        rd = row["recorded_at"]; rdd = rd[:10] if isinstance(rd, str) else str(rd)[:10]
        vals = [ri - 1, rdd, gn, pn, ic, mb, it, row["quantity"], row["user_name"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val); c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T

    for col, w in [(1, 6), (2, 12), (3, 14), (4, 30), (5, 10), (6, 30), (7, 8), (8, 8), (9, 12)]:
        ws.column_dimensions[_cl(col)].width = w
    ws.freeze_panes = "A2"


# ═══════════════════  Sheet 5: 用户统计 ═══════════════════
def _build_user_stats(ws, db, start_s, end_s, group_id):
    ws.sheet_properties.tabColor = "9C27B0"
    if start_s: ref = datetime.strptime(start_s[:10], "%Y-%m-%d")
    else: ref = datetime.now()
    d_start, d_end = _month_bounds(ref)

    rc = ["wr.deleted_at IS NULL"]; rp = []
    if start_s: rc.append("wr.recorded_at >= ?"); rp.append(start_s)
    if end_s: rc.append("wr.recorded_at <= ?"); rp.append(end_s[:10] + "T23:59:59")
    if group_id: rc.append("pg.id = ?"); rp.append(group_id)
    rows = db.execute(
        f"SELECT wr.user_name, pg.name AS group_name, p.name AS project_name, SUM(wr.quantity) AS qty, COUNT(*) AS cnt FROM work_records wr JOIN projects p ON wr.project_id = p.id JOIN project_groups pg ON p.group_id = pg.id WHERE {' AND '.join(rc)} GROUP BY wr.user_name, p.id ORDER BY wr.user_name, pg.sort_order, p.sort_order", rp).fetchall()

    CB, CC, CD, CE, CF, CQ = 2, 3, 4, 5, 6, 7
    for col, label in [(CB, "用户名"), (CC, "使用实验室"), (CD, "项目代号"), (CE, "液相仪器"), (CF, "检测方法"), (CQ, "数量")]:
        c = ws.cell(row=2, column=col, value=label); c.font = FONT_H; c.alignment = ALIGN_C; c.border = BORDER_T

    r = 3
    for row in rows:
        un, gn, pn = row["user_name"], row["group_name"], row["project_name"]
        pc = _extract_project_code(pn); mb, ic, _ = _parse_instrument(pn)
        fn = get_method_full_name(gn, pn) or mb
        for col, val in [(CB, un), (CC, gn), (CD, pc), (CE, ic), (CF, fn), (CQ, row["qty"] or None)]:
            c = ws.cell(row=r, column=col, value=val); c.font = FONT_D; c.alignment = ALIGN_C; c.border = BORDER_T
        r += 1

    ws.column_dimensions['A'].width = 3
    for col, w in [(2, 12), (3, 18), (4, 12), (5, 10), (6, 43), (7, 10)]:
        ws.column_dimensions[_cl(col)].width = w
    ws.freeze_panes = "A2"


@router.get("/excel")
def export_excel(start: str = Query(None), end: str = Query(None), group_id: int = Query(None), db=Depends(get_db)):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "月-汇总"; ws1.sheet_properties.tabColor = "1976D2"
    month_label = _build_monthly_summary(ws1, db, start, end, group_id)
    ws2 = wb.create_sheet(title="每日工作量"); _build_daily_work(ws2, db, start, end, group_id)
    ws3 = wb.create_sheet(title="每周工作量"); _build_weekly_work(ws3, db, start, end, group_id)
    ws4 = wb.create_sheet(title="原始记录"); _build_raw_records(ws4, db, start, end, group_id)
    ws5 = wb.create_sheet(title="用户统计"); _build_user_stats(ws5, db, start, end, group_id)
    output = BytesIO(); wb.save(output); output.seek(0)
    filename = quote(f"工作量统计_{month_label}.xlsx")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
